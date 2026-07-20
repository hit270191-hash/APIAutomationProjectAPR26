import json
import os
from openpyxl import load_workbook


#json data provider
def read_json_data(filename):
    with open(filename) as f:
        data = json.load(f)
        return [(item,) for item in data]

def read_excel_data(filename, sheet_name):
    # opens xl file
    wb= load_workbook(filename)
    # selects the specified sheet
    sheet= wb[sheet_name]
    # reads first row (headings)
    header= [cell.value for cell in sheet[1]]

    print(header)

    data=[]
    for row in sheet.iter_rows(min_row= 2, values_only= True):
        if any(row):
            data.append(dict(zip(header,row)))

    return data

## Practice: Perform DDT using CSV file
